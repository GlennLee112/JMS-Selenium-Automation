# Data processing modules
import JMS
from datetime import datetime
import sys

# Openpyxl modules: data writing to xlsx files
import os
import openpyxl as opx
from openpyxl import Workbook
from openpyxl.styles import NamedStyle

# Pasting modules
import pyperclip

# Selenium modules
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException, TimeoutException, ElementNotInteractableException, ElementClickInterceptedException

# Selenium Drivers
#_________________________________________________________#
driver = JMS.Driver()

#_________________________________________________________#

# Open workbook
#_________________________________________________________#
path = os.getcwd()
xlsx_path = os.path.join(path,"test.xlsx") # To replace the file name with the mode of month on dataframe read

xlsx_file = Workbook()
worksheet_list = xlsx_file.get_sheet_names()
print(worksheet_list)
worksheet = xlsx_file.get_sheet_by_name('Sheet')

# cell formating solution, reference:
# https://dev.to/abbazs/how-to-change-date-format-in-excel-using-openpyxl-4fnm
date_format = NamedStyle(name="date1", number_format="YYYY-MM-DD")
row_to_use = 2

#_________________________________________________________#
# Testing AWB list
# Test_AWB = '688575345125'
# Overnight_Scan_2_AWB = '688583059465'
# Multi_overnight_scan = 'Wait for Renu to provide'
# No_overnight_scan = '650689924152'

#_________________________________________________________#
# Departure section:
#_________________________________________________________#
try:
    # XPATH for departure
    Departure_xpath = "//tr[.//td[contains(., 'Departure')] and .//td[contains(., 'CDC TIONGEMAS 301')]]"

    # Find the row elements
    Departure_element = driver.find_elements(By.XPATH, Departure_xpath)
    print(len(Departure_element)) # Test if elements are found (i.e.: not zero); disable with '#' if testing not needed

    # Extract the time value from the second column in that row
    time_xpath = "./td[3]//span"
    Departure_time = datetime.strptime((Departure_element[0].find_element(By.XPATH, time_xpath).text),
                                       '%Y-%m-%d %H:%M:%S')
    Departure_time = Departure_time.strftime('%Y-%m-%d')

    print('Departure time is:', Departure_time)

    cell = worksheet.cell(row=row_to_use,column=4)
    cell.value = Departure_time
    cell.style = date_format

except IndexError:
    print('Departure is passed')
    pass

#_________________________________________________________#

# Arrival section:
#_________________________________________________________#
try:
    Arrival_xpath = f"//tr[.//td[contains(., 'Arrival')] and .//td[contains(., 'CDC TIONGEMAS 301')]]"

    # Find the row element
    Arrival_elements = driver.find_elements(By.XPATH, Arrival_xpath)

    print(len(Arrival_elements)) # Test if elements are found (i.e.: not zero); disable with '#' if testing not needed

    Arrival_index = 0
    # Extract the time value from the second column in that row
    time_xpath = "./td[3]//span"
    # Convert text to datetime
    Arrival_1 = datetime.strptime((Arrival_elements[Arrival_index].find_element(By.XPATH, time_xpath).text),
                                  '%Y-%m-%d %H:%M:%S')
    # Strip datetime to date only
    Arrival_1 = Arrival_1.strftime('%Y-%m-%d')

    print('Arrival Time is:', Arrival_1)
    #
    cell = worksheet.cell(row=row_to_use,column=5)
    cell.value = Arrival_1
    cell.style = date_format

except IndexError:
    print('Arrival is passed')
    pass

#_________________________________________________________#
# Return Registration
#_________________________________________________________#

# Return Registration date
try:
    Return_Reg_xpath = ("//div[@class='el-table__body-wrapper is-scrolling-none']//tbody//"
                       "tr[.//td[contains(., 'Return Registration')]and .//td[contains(., 'CDC TIONGEMAS 301')]]")

    Return_Reg_elements = driver.find_elements(By.XPATH, Return_Reg_xpath)
    time_xpath = "./td[3]//span"

    Return_Reg = datetime.strptime((Return_Reg_elements[0].find_element(By.XPATH, time_xpath).text),
                                  '%Y-%m-%d %H:%M:%S')
    Return_Reg = Return_Reg.strftime('%Y-%m-%d')

    print('Return Registration Time is:', Return_Reg)

    cell = worksheet.cell(row=row_to_use,column=5)
    cell.value = Return_Reg
    cell.style = date_format

except IndexError:
    print('Return Reg passed')
    pass
#_________________________________________________________#
# Overnight section:
#_________________________________________________________#

# Overnight date
# 1st Overnight
Overnight_xpath = ("//div[@class='el-table__body-wrapper is-scrolling-none']//tbody//tr[.//td[contains(., 'Overnight')]"
                   " and .//td[contains(., 'CDC TIONGEMAS 301')]]")
# Success: Future tip - work HTML from ascending order (ie: from CDC to the front)

# Overnight Elements find & Time XPATH
Overnight_elements = driver.find_elements(By.XPATH, Overnight_xpath)
time_xpath = "./td[3]//span"

# Counter & index Setup
Overnight_counter = 0
Overnight_index = 0
Overnight_variable_no = 1


# Empty dict. for dynamic Overnight Variables
Overnight_var = {}

# Loop for 3 times -- 3 Overnight Scan
while Overnight_counter < 3:
    # For first loop

    overnight_column = 7 + (Overnight_variable_no - 1)

    if Overnight_counter == 0: # counter == 0 is the first loop
        # Automatically add 1 to counter
        #Overnight_counter += 1
        try:
            Overnight_var[f'Overnight_{Overnight_variable_no}'] = datetime.strptime(
                (Overnight_elements[Overnight_index].find_element
                 (By.XPATH, time_xpath).text), '%Y-%m-%d %H:%M:%S')
            Overnight_var[f'Overnight_{Overnight_variable_no}'] = (Overnight_var[f'Overnight_{Overnight_variable_no}']
                                                                   .strftime('%Y-%m-%d'))

        # No Exception as first overnight time is assumed to always be present
        except IndexError:
            print('No overnight scan is done') # Debugging print; disable or enable as needed (by adding '#')
            break
        print('Overnight ', Overnight_variable_no, ' is done on ', Overnight_var[f'Overnight_{Overnight_variable_no}'])

        cell = worksheet.cell(row=row_to_use, column=overnight_column)
        cell.value = Overnight_var[f'Overnight_{Overnight_variable_no}']
        cell.style = date_format
        # Automatically add 1 to counter and variable no.
        Overnight_variable_no += 1
        Overnight_counter += 1

    # For subsequent Loop after first -- need to account for same date to the previous overnight scan
    elif Overnight_counter > 0:

        # Automatically add 1 to counter
        #Overnight_counter += 1
        try:
            # attempting in finding overnight date
            Overnight_var[f'Overnight_{Overnight_variable_no}'] = datetime.strptime(
                (Overnight_elements[Overnight_index].find_element
                 (By.XPATH, time_xpath).text), '%Y-%m-%d %H:%M:%S')
            Overnight_var[f'Overnight_{Overnight_variable_no}'] = Overnight_var[f'Overnight_{Overnight_variable_no}'].strftime('%Y-%m-%d')

            # Account for if the previous date is same as the current date
            if Overnight_var[f'Overnight_{Overnight_variable_no}'] == Overnight_var[f'Overnight_{Overnight_variable_no - 1}']:

                # Loop to next element until Overnight scan with a different date is found
                while Overnight_var[f'Overnight_{Overnight_variable_no}'] == Overnight_var[f'Overnight_{Overnight_variable_no - 1}']:

                    Overnight_index += 1
                    # Find element of next overnight scan
                    Overnight_var[f'Overnight_{Overnight_variable_no}'] = datetime.strptime(
                        (Overnight_elements[Overnight_index].find_element
                         (By.XPATH, time_xpath).text), '%Y-%m-%d %H:%M:%S')

                    Overnight_var[f'Overnight_{Overnight_variable_no}'] = (Overnight_var[f'Overnight_{Overnight_variable_no}']
                                                                           .strftime('%Y-%m-%d'))
            else:
                Overnight_index += 1

        except IndexError:
            print('No more overnight scan')
            break
        print('Overnight ', Overnight_variable_no, ' is done on ', Overnight_var[f'Overnight_{Overnight_variable_no}'])
        cell = worksheet.cell(row=row_to_use, column=overnight_column)
        cell.value = Overnight_var[f'Overnight_{Overnight_variable_no}']
        cell.style = date_format

        # Automatically add 1 to counter and variable no.
        Overnight_variable_no += 1
        Overnight_counter += 1

# xlsx_file.save(xlsx_path) # save the file once done writing