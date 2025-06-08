# Data processing modules
import os
import glob
import pandas as pd
import numpy as np
import time
from datetime import datetime
import calendar

# Pasting modules
import pyperclip

# JMS and modules
import JMS

# Selenium modules
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException

# Openpyxl modules: data writing to xlsx files
import openpyxl as opx
from openpyxl import Workbook
from openpyxl.styles import NamedStyle

# Revision aim:
# 1. Let selenium take the last arrival date
# 2. Have overnight taking into account of arrival date

driver = JMS.Driver()

# _________________________________________________________ #
# Arrival section:
# _________________________________________________________ #

# Empty arrival to be assigned for used outside of try block
# Reference: https://stackoverflow.com/questions/25666853/how-to-make-a-variable-inside-a-try-except-block-public
Arrival_test = ''

try:
    Arrival_xpath = (f"//div[@class='el-table__body-wrapper is-scrolling-none']//tr[.//td[contains(., 'Arrival')] "
                     f"and .//td[contains(., 'CDC LARKIN 334')]]")

    # Find the row element
    Arrival_elements = driver.find_elements(By.XPATH, Arrival_xpath)

    # Select elements to extract from; the last element
    Arrival_index = len(Arrival_elements) - 1
    # Extract the time value from the second column in that row
    time_xpath = "./td[3]//span"
    # Convert text to datetime
    Arrival_1 = datetime.strptime((Arrival_elements[Arrival_index].find_element(By.XPATH, time_xpath).text),
                                  '%Y-%m-%d %H:%M:%S')

    Arrival_test = Arrival_1.date()

    # Strip datetime to date only
    Arrival_1 = Arrival_1.strftime('%Y-%m-%d %H:%M:%S')

    print('Arrival Time is:', Arrival_1)
    #
    # cell = worksheet.cell(row=row_to_use, column=5)
    # cell.value = Arrival_1
    # cell.style = date_format

except (NoSuchElementException, StaleElementReferenceException, IndexError,) as e:
    #print('Arrival is passed')
    pass

# _________________________________________________________ #
# Overnight Section
# _________________________________________________________ #

Overnight_xpath = (
    f"//div[@class='el-table__body-wrapper is-scrolling-none']//tbody//tr[.//td[contains(., 'Overnight Scan')] "
    f"and .//td[contains(., 'CDC LARKIN 334')]]")
# Success: Future tip - work HTML from ascending order (ie: from CDC to the front)

# Overnight Elements find & Time XPATH
Overnight_elements = driver.find_elements(By.XPATH, Overnight_xpath)
time_xpath = "./td[3]//span"

# Overnight Counter & index Setup
Overnight_counter = 0
Overnight_index = 0
Overnight_variable_no = 1

# Empty dict. for dynamic Overnight Variables
Overnight_var = {}

# Overnight index correction, compensate variable until the
# test overnight
Overnight_index_add = True

while Overnight_index_add:
    try:
        test_overnight = datetime.strptime(
            (Overnight_elements[Overnight_index].find_element
             (By.XPATH, time_xpath).text), '%Y-%m-%d %H:%M:%S').date()
    except (NoSuchElementException, IndexError,) as e:
        print('Error in Overnight date, ending the loop earlier')
        break

    if test_overnight < Arrival_test:
        Overnight_index += 1
        print('Overnight is smaller than arrival; adding one to index')
        Overnight_index_add = True
    else:
        print(f'Overnight is now normal; stopped adding index and index at :', Overnight_index)
        Overnight_index_add = False

    # Overnight_index += 1
    # print('Add 1 overnight_index')

# Overnight main loop
while Overnight_counter < 5:
    # For first loop

    overnight_column = 7 + (Overnight_variable_no - 1)

    if Overnight_counter == 0:  # counter == 0 is the first loop
        # Automatically add 1 to counter
        # Overnight_counter += 1
        try:
            Overnight_var[f'Overnight_{Overnight_variable_no}'] = datetime.strptime(
                (Overnight_elements[Overnight_index].find_element
                 (By.XPATH, time_xpath).text), '%Y-%m-%d %H:%M:%S')

            Overnight_var[f'Overnight_{Overnight_variable_no}'] = (
                Overnight_var[f'Overnight_{Overnight_variable_no}']
                .strftime('%Y-%m-%d'))

        # except Overnight_var[f'Overnight_{Overnight_variable_no}'] < Arrival_1:
        #     while

        # No Exception as first overnight time is assumed to always be present
        except (NoSuchElementException, StaleElementReferenceException, IndexError,) as e:
            # print('No overnight scan is done')  # Debugging print; disable or enable as needed (by adding '#')
            break

        print('Overnight ', Overnight_variable_no, ' is done on ', Overnight_var[f'Overnight_{Overnight_variable_no}'])
        # For test printing overnight variable for validation; add or remove "#" to as needed to run / not run test.

        # cell = worksheet.cell(row=row_to_use, column=overnight_column)
        # cell.value = Overnight_var[f'Overnight_{Overnight_variable_no}']
        # cell.style = date_format
        # Automatically add 1 to counter and variable no.
        Overnight_variable_no += 1
        Overnight_counter += 1

    # For subsequent Loop after first -- need to account for same date to the previous overnight scan
    elif Overnight_counter > 0:

        # Automatically add 1 to counter
        # Overnight_counter += 1
        try:
            # attempting in finding overnight date
            Overnight_var[f'Overnight_{Overnight_variable_no}'] = datetime.strptime(
                (Overnight_elements[Overnight_index].find_element
                 (By.XPATH, time_xpath).text), '%Y-%m-%d %H:%M:%S')
            Overnight_var[f'Overnight_{Overnight_variable_no}'] = Overnight_var[
                f'Overnight_{Overnight_variable_no}'].strftime('%Y-%m-%d')

            # Account for if the previous date is same as the current date
            if Overnight_var[f'Overnight_{Overnight_variable_no}'] == Overnight_var[
                f'Overnight_{Overnight_variable_no - 1}']:

                # Loop to next element until Overnight scan with a different date is found
                while (Overnight_var[f'Overnight_{Overnight_variable_no}'] ==
                       Overnight_var[f'Overnight_{Overnight_variable_no - 1}']):
                    Overnight_index += 1
                    # Find element of next overnight scan
                    Overnight_var[f'Overnight_{Overnight_variable_no}'] = datetime.strptime(
                        (Overnight_elements[Overnight_index].find_element
                         (By.XPATH, time_xpath).text), '%Y-%m-%d %H:%M:%S')

                    Overnight_var[f'Overnight_{Overnight_variable_no}'] = (
                        Overnight_var[f'Overnight_{Overnight_variable_no}']
                        .strftime('%Y-%m-%d'))
            else:
                Overnight_index += 1

        except (NoSuchElementException, StaleElementReferenceException, IndexError,) as e:
            # print('No more overnight scan')
            break
        print('Overnight ', Overnight_variable_no, ' is done on ', Overnight_var[f'Overnight_{Overnight_variable_no}'])
        # cell = worksheet.cell(row=row_to_use, column=overnight_column)
        # cell.value = Overnight_var[f'Overnight_{Overnight_variable_no}']
        # cell.style = date_format

        # Automatically add 1 to counter and variable no.
        Overnight_variable_no += 1
        Overnight_counter += 1