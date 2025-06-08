# Data processing modules
import os
import glob
import pandas as pd
import numpy as np
import time
from datetime import datetime
import calendar

# Pasting modules
import pyperclip

# JMS and modules
import JMS

# Selenium modules
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException

# Openpyxl modules: data writing to xlsx files
import openpyxl as opx
from openpyxl import Workbook
from openpyxl.styles import NamedStyle

#_________________________________________________________
# 0. Drivers & Paths
#_________________________________________________________
# Driver
driver = JMS.Driver()

# Paths
path = os.getcwd()
AWB_folder = glob.glob(os.path.join(path,"AWB","*.xlsx"))
DP_folder = os.path.join(path,'DP List')
DP_list = glob.glob(os.path.join(DP_folder,"*.xlsx"))

# Folders to nuke
Folder_Nuke = [DP_folder]

#_________________________________________________________
# 1. Load AWB xlsx file to read
#_________________________________________________________

## Empty list for append
AWB = []

# DF merging reference:
# https://saturncloud.io/blog/how-to-merge-multiple-sheets-from-multiple-excel-workbooks-into-a-single-pandas-dataframe/

# Read AWB Folder for all AWB xlsx
for A in AWB_folder:
    # print(A) #Remove '#' to debug / test
    df = pd.read_excel(A)
    AWB.append(df)

# Concat AWB dfs into one and reset index after appending files
AWB = pd.concat(AWB).reset_index()

#_________________________________________________________#
# 2. DP Lists download from JMS if needed
#_________________________________________________________#

# Use the dataframes loaded from step 1 to find the mode of date for extracting required info
AWB_date = AWB['Appeal Date'].mode()

# Excel Datetime conversion, reference:
# https://stackoverflow.com/questions/31359150/convert-date-from-excel-in-number-format-to-date-format-python
AWB_date = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + AWB_date[0] - 2)

# Get year and month of date to use for JMS range
AWB_Mode_year = AWB_date.year
AWB_Mode_month = AWB_date.month

# First and last day of the month, reference:
# https://stackoverflow.com/questions/70530970/how-to-get-the-first-and-last-date-of-the-month-with-year-and-month-in-python
first, last = calendar.monthrange(AWB_Mode_year, AWB_Mode_month)

Start_Date_D_JMS = str(datetime(AWB_Mode_year, AWB_Mode_month, 1))
End_Date_D_JMS = str(datetime(AWB_Mode_year, AWB_Mode_month, last))

    # Logic date - Download DP name list if folder is empty
DP_list = glob.glob(os.path.join(DP_folder,"*.xlsx"))

if len(DP_list) == 0:
    JMS.Delivery_Summary_Lump(driver, Start_Date_D_JMS, End_Date_D_JMS, DownLoc=DP_folder)
else:
    pass
#_________________________________________________________#
# 3. Merge AWB df with DP name list
#_________________________________________________________#

    # DP folder to read DP names from
DP_list_df = pd.read_excel(DP_list[0])
DP_list_df_Merge = DP_list_df[['DP No. | Delivery','Delivery DP']]

    # Merge the DP List with AWB dataframe
New_AWB = pd.merge(AWB, DP_list_df_Merge, left_on='DP', right_on='DP No. | Delivery').drop('DP No. | Delivery',axis=1)
#New_AWB['AWB'] = New_AWB['AWB'].astype(str)

    # Disable to test output AWB to xlsx
    #New_AWB.to_excel(os.path.join(path,"New AWB.xlsx"), index=False)

#_________________________________________________________#
# 4. Loop, JMS, and write to excel
#_________________________________________________________#

# Multi layer loop, please note of all the layer for clarity

# Counter(s) Setup
Counter = 0 # Used by layer 0 to track loop progress and end when needed
DF_Len = len(New_AWB) # As the stopped for layer 0 loop

# Starting and ending ILoc setup
ILoc_Start = 0
ILoc_End = 999

# Row to use temporary holder
row_to_use = 2

# open workbook
xlsx_path = os.path.join(path,'Output',"test.xlsx") # To replace the file name with the mode of month on dataframe read
xlsx_file = Workbook() # open empty workbook
worksheet_list = xlsx_file.get_sheet_names()
#print(worksheet_list)
worksheet = xlsx_file.get_sheet_by_name('Sheet')

# write header to empty workbook
Header_list = ['AWB', 'DP No.', 'Timeliness', 'Departure', 'Arrival', 'Return Registration', 'Overnight 1',
               'Overnight 2', 'Overnight 3', 'Overnight 4', 'Overnight 5', 'Overnight 6', 'Signature', 'ATS']

Header_column = 1

for Header_items in Header_list:
    cell = worksheet.cell(row=1, column=Header_column)
    cell.value = Header_items
    Header_column += 1

# JMS Parcel Tracking Navigation
driver.get('https://jms.jtexpress.my/indexSub')
# XPATH for parcel tracking button
Parcel_Tracking_xpath = "//div[@class='header-right']//div[contains(@class, 'hr-search') and contains(text(), 'Parcel Tracking')]"
# Conditionally wait for button element to appear
Parcel_Tracking_elem = WebDriverWait(driver, 60).until(
    EC.presence_of_element_located((By.XPATH, Parcel_Tracking_xpath)))
# Click on JMS Parcel Tracking button
Parcel_Tracking_elem.click()

# Layer 1 Loop: AWB dataframe ILoc
while Counter < DF_Len:

    # Select the AWB dataframe range to use useing ILoc var.
    Temp_DF = New_AWB.iloc[ILoc_Start:ILoc_End]  # Will
    txt_df = Temp_DF['AWB']

    ## Save AWBs to tmp text for pasting to search box
    # file path for txt. file holding AWB no.
    tmp_text_file = os.path.join(path, "tmp.txt")

    # Saving
    np.savetxt(tmp_text_file, txt_df.values, fmt='%d')

    tracking_search_xpath_1 = "//div[@class='vue-input-tag-wrapper']/input"
    tracking_search_xpath_2 = "//div[@class='vue-input-tag-wrapper vue-input-tag-wrapper--active']/input"

    try:
        Tracking_boxes = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, tracking_search_xpath_1)))
    except NoSuchElementException:
        Tracking_boxes = WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.XPATH, tracking_search_xpath_2)))

    Tracking_boxes.click()

    ## Option 2: clipboard (pyperclip)
    ## Open clipboard with saved text for reading and then pasting
    with open(tmp_text_file, 'r') as txt_file:
        content = txt_file.read()
    # To clipboard
    pyperclip.copy(content)
    # crtl + v (pasting) to JMS box
    Tracking_boxes.send_keys(Keys.CONTROL + 'v')
    # Hard pause for 0.5 Second
    time.sleep(0.5)
    # Click 'Inquire'
    WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH,
                                        "//button[contains(@class, 'el-button first-btn el-button--default el-button--small') and .//i[contains(@class, 'el-icon-search')] and .//span[text()='Inquire']]"))
    ).click()

    time.sleep(0.2)

    # Counter, empty var., and format solution for layer 2
    ## Counter
    Button_Counter = 0
    ## Empty Var
    order_no_hold = ''
    ## cell formating solution, reference:
    ## https://dev.to/abbazs/how-to-change-date-format-in-excel-using-openpyxl-4fnm
    date_format = NamedStyle(name="date1", number_format="YYYY-MM-DD HH:MM:SS")

    for index, row in Temp_DF.iterrows():

        tracking_list_xpath = f"//div[@class='tk_list_child_item' and contains(text(),'{row['AWB']}')]"
        tracking_list_xpath_active = f"//div[@class='tk_list_child_item active' and contains(text(),'{row['AWB']}')]"
        # Correct_Date = pd.TimedeltaIndex(row['Appeal Date'], unit='d') + dt.datetime(1899, 12, 30)

        AWB_Button = driver.find_elements(By.XPATH, tracking_list_xpath)
        # print(len(AWB_Button))
        ## Logic gate in case tracking_no_xpath returns no result
        if len(AWB_Button) == 0:
            AWB_Button = driver.find_elements(By.XPATH, tracking_list_xpath_active)
        else:
            pass
        # print(row['AWB'])  # Disable when not testing
        Button_Counter += 1
        # print(Button_Counter)  # Disable when not testing
        AWB_Button[len(AWB_Button) - 1].click()
        # print(row['Delivery DP'])

        # _________________________________________________________#
        # AWB Section:
        # _________________________________________________________#

        # Write AWB
        cell = worksheet.cell(row=row_to_use, column=1)
        cell.value = str(row['AWB'])
        print([row['AWB']])

        # _________________________________________________________#
        # DP Section:
        # _________________________________________________________#

        # Write DP
        # Write AWB
        cell = worksheet.cell(row=row_to_use, column=2)
        cell.value = str(row['DP'])

        # _________________________________________________________#
        # DP Section:
        # _________________________________________________________#

        # Write DP
        # Write AWB
        cell = worksheet.cell(row=row_to_use, column=3)
        cell.value = str(row['Timeliness'])


        # _________________________________________________________#
        # Order No. section:
        # _________________________________________________________#
        # Temporary hold order no. for comparison and validation
        # Find the element and extract the number

        # Find the element and extract the number
        order_no_xpath = "//div[@class='item-msg']//div[@class='tt' and text()='Order No.']/following-sibling::div[@class='des']"
        order_no_element = driver.find_element(By.XPATH, order_no_xpath)
        order_no_value = order_no_element.text.strip()

        # Hold the loop while page order no is still the same as hold order no.
        while order_no_hold == order_no_value:
            time.sleep(0.01)
            order_no_element = driver.find_element(By.XPATH, order_no_xpath)
            order_no_value = order_no_element.text.strip()

        # Overwrite order no hold with the new order no for comparison purpose
        order_no_element = driver.find_element(By.XPATH, order_no_xpath)
        order_no_hold = order_no_element.text.strip()

        # Start navigate and writing data
        # _________________________________________________________#
        # Departure section:
        # _________________________________________________________#

        # Improvement needed: To account for appeal date and the timeliness, I.e.: distinguish departure scan that is
        # earlier than the appeal date and timeliness.

        Departure_test = ''
        Departure_test_failed = False
        while True:
            try:
                # Composite partial text XPATH reference: https://stackoverflow.com/questions/54498277/selenium-use-multiple-strings-in-find-element-by-partial-link-text
                Departure_xpath = (f"//div[@class='el-table__body-wrapper is-scrolling-none']"
                                   f"//tr[.//td[contains(., 'Departure')] "
                                   f"and .//td[contains(., '{row['Delivery DP']}') and contains(., 'and sent to [ ')]]")

                # Find the row elements
                Departure_element = driver.find_elements(By.XPATH, Departure_xpath)


                #ignored_exceptions = (StaleElementReferenceException,)
                #Departure_element = (WebDriverWait(driver,10,ignored_exceptions=ignored_exceptions)
                #                     .until(EC.presence_of_all_elements_located((By.XPATH, Departure_xpath))))

                #print(len(Departure_element))  # Test if elements are found (i.e.: not zero); disable with '#' if testing not needed

                # Extract the time value from the second column in that row
                time_xpath = "./td[3]//span"
                Departure_time = datetime.strptime((Departure_element[0].find_element(By.XPATH, time_xpath).text),
                                                   '%Y-%m-%d %H:%M:%S')

                # Assign test variable for testing in overnight
                Departure_test = Departure_time.date()

                Departure_time = Departure_time.strftime('%Y-%m-%d %H:%M:%S')

                print('Departure time is:', Departure_time)
                print('Departure test time is:', Departure_test)

                cell = worksheet.cell(row=row_to_use, column=4)
                cell.value = Departure_time
                cell.style = date_format
                break
            except StaleElementReferenceException:
                print('Stale Departure, repeating')
                break

            except (NoSuchElementException, IndexError,) as e:
                print('Departure is passed')
                Departure_test_failed = True
                break

        # _________________________________________________________#
        # Arrival section:
        # Find the arrival datetime
        # _________________________________________________________#
        # empty variable; arrival_test variable to be used for comparing with overnight date

        # Improvement required: using reducing loop, find the earliest arrival scan that is after departure time

        Arrival_test = ''
        Arrival_test_failed = False

        try:
            Arrival_xpath = (f"//div[@class='el-table__body-wrapper is-scrolling-none']//tr[.//td[contains(., 'Arrival')] "
                             f"and .//td[contains(., '{row['Delivery DP']}') and contains(., 'The parcel arrived at')]]")

            # Find the row element
            Arrival_elements = driver.find_elements(By.XPATH, Arrival_xpath)
            #ignored_exceptions = (StaleElementReferenceException,)
            #Arrival_elements = (WebDriverWait(driver, 10,ignored_exceptions=ignored_exceptions)
            #                    .until(EC.presence_of_all_elements_located((By.XPATH, Arrival_xpath))))

            #print(len(Arrival_elements))  # Test if elements are found (i.e.: not zero); disable with '#' if testing not needed

            Arrival_index = len(Arrival_elements) - 1
            # Extract the time value from the second column in that row
            time_xpath = "./td[3]//span"
            # Convert text to datetime
            Arrival_1 = datetime.strptime((Arrival_elements[Arrival_index].find_element(By.XPATH, time_xpath).text),
                                          '%Y-%m-%d %H:%M:%S')
            Arrival_test = Arrival_1.date()

            # Strip datetime to date only
            Arrival_1 = Arrival_1.strftime('%Y-%m-%d %H:%M:%S')

            print('Arrival Time is:', Arrival_1)
            #
            cell = worksheet.cell(row=row_to_use, column=5)
            cell.value = Arrival_1
            cell.style = date_format

        except (NoSuchElementException, StaleElementReferenceException, IndexError,) as e:
            Arrival_test_failed = True
            #print('Arrival is passed')
            pass

        # _________________________________________________________#
        # Return Registration
        # _________________________________________________________#

        # Return Registration date
        try:
            Return_Reg_xpath = (f"//div[@class='el-table__body-wrapper is-scrolling-none']//tbody//"
                                f"tr[.//td[contains(., 'Return Registration')]and .//td[contains(., '{row['Delivery DP']}')]]")

            #ignored_exceptions = (StaleElementReferenceException,)

            Return_Reg_elements = driver.find_elements(By.XPATH, Return_Reg_xpath)
            #Return_Reg_elements = (WebDriverWait(driver, 30,ignored_exceptions=ignored_exceptions)
            #                       .until(EC.presence_of_all_elements_located((By.XPATH, Return_Reg_xpath))))

            time_xpath = "./td[3]//span"

            Return_Reg = datetime.strptime((Return_Reg_elements[0].find_element(By.XPATH, time_xpath).text),
                                           '%Y-%m-%d %H:%M:%S')

            Return_Reg = Return_Reg.strftime('%Y-%m-%d')

            #print('Return Registration Time is:', Return_Reg)

            cell = worksheet.cell(row=row_to_use, column=6)
            cell.value = Return_Reg
            cell.style = date_format

        except (NoSuchElementException, StaleElementReferenceException,IndexError,) as e:
            #print('Return Reg passed')
            pass

        # _________________________________________________________#
        # Overnight section:
        # _________________________________________________________#
        # Overnight date
        # 1st Overnight
        Overnight_xpath = (
            f"//div[@class='el-table__body-wrapper is-scrolling-none']//tbody//tr[.//td[contains(., 'Overnight Scan')] "
            f"and .//td[contains(., '{row['Delivery DP']}')]]")
        # Success: Future tip - work HTML from ascending order (ie: from CDC to the front)

        # Overnight Elements find & Time XPATH
        Overnight_elements = driver.find_elements(By.XPATH, Overnight_xpath)
        time_xpath = "./td[3]//span"

        # Overnight Counter & index Setup
        Overnight_counter = 0
        Overnight_index = 0
        Overnight_variable_no = 1

        # Empty dict. for dynamic Overnight Variables
        Overnight_var = {}

        # Overnight index correction, compensate variable until the overnight is not earlier than arrival
        # test overnight
        Overnight_index_add = True
        while Overnight_index_add:
            # Assign test_overnight date value
            try:
                test_overnight = datetime.strptime(
                    (Overnight_elements[Overnight_index].find_element
                     (By.XPATH, time_xpath).text), '%Y-%m-%d %H:%M:%S').date()

                if ((Arrival_test_failed == False and Departure_test_failed) or
                        (Arrival_test_failed == False and Departure_test_failed == False)):
                    if test_overnight < Arrival_test:
                        Overnight_index += 1
                        print('Overnight is smaller than arrival; adding one to index')
                        Overnight_index_add = True
                    else:
                        print(f'Overnight is now normal; stopped adding index and index at :', Overnight_index)
                        Overnight_index_add = False

                elif Arrival_test_failed and Departure_test_failed == False:
                    if test_overnight < Departure_test:
                        Overnight_index += 1
                        print('Overnight is smaller than arrival; adding one to index')
                        Overnight_index_add = True
                    else:
                        print(f'Overnight is now normal; stopped adding index and index at :', Overnight_index)
                        Overnight_index_add = False
                else:
                    print('Both Arrival test and departure test not available')
                    break
            except (NoSuchElementException, IndexError,) as e:
                print('Error in Overnight date, ending the loop earlier')
                break

        # Loop for 3 times -- 3 Overnight Scan
        while Overnight_counter < 6:
            # For first loop

            overnight_column = 7 + (Overnight_variable_no - 1)

            if Overnight_counter == 0:  # counter == 0 is the first loop
                # Automatically add 1 to counter
                # Overnight_counter += 1
                try:
                    Overnight_var[f'Overnight_{Overnight_variable_no}'] = datetime.strptime(
                        (Overnight_elements[Overnight_index].find_element
                         (By.XPATH, time_xpath).text), '%Y-%m-%d %H:%M:%S')
                    Overnight_var[f'Overnight_{Overnight_variable_no}'] = (
                        Overnight_var[f'Overnight_{Overnight_variable_no}']
                        .strftime('%Y-%m-%d'))

                # No Exception as first overnight time is assumed to always be present
                except (NoSuchElementException, StaleElementReferenceException, IndexError,) as e:
                    #print('No overnight scan is done')  # Debugging print; disable or enable as needed (by adding '#')
                    break
                #print('Overnight ', Overnight_variable_no, ' is done on ', Overnight_var[f'Overnight_{Overnight_variable_no}'])

                cell = worksheet.cell(row=row_to_use, column=overnight_column)
                cell.value = Overnight_var[f'Overnight_{Overnight_variable_no}']
                cell.style = date_format
                # Automatically add 1 to counter and variable no.
                Overnight_variable_no += 1
                Overnight_counter += 1

            # For subsequent Loop after first -- need to account for same date to the previous overnight scan
            elif Overnight_counter > 0:

                # Automatically add 1 to counter
                # Overnight_counter += 1
                try:
                    # attempting in finding overnight date
                    Overnight_var[f'Overnight_{Overnight_variable_no}'] = datetime.strptime(
                        (Overnight_elements[Overnight_index].find_element
                         (By.XPATH, time_xpath).text), '%Y-%m-%d %H:%M:%S')
                    Overnight_var[f'Overnight_{Overnight_variable_no}'] = Overnight_var[
                        f'Overnight_{Overnight_variable_no}'].strftime('%Y-%m-%d')

                    # Account for if the previous date is same as the current date
                    if Overnight_var[f'Overnight_{Overnight_variable_no}'] == Overnight_var[
                        f'Overnight_{Overnight_variable_no - 1}']:

                        # Loop to next element until Overnight scan with a different date is found
                        while Overnight_var[f'Overnight_{Overnight_variable_no}'] == Overnight_var[f'Overnight_{Overnight_variable_no - 1}']:
                            Overnight_index += 1
                            # Find element of next overnight scan
                            Overnight_var[f'Overnight_{Overnight_variable_no}'] = datetime.strptime(
                                (Overnight_elements[Overnight_index].find_element
                                 (By.XPATH, time_xpath).text), '%Y-%m-%d %H:%M:%S')

                            Overnight_var[f'Overnight_{Overnight_variable_no}'] = (
                                Overnight_var[f'Overnight_{Overnight_variable_no}']
                                .strftime('%Y-%m-%d'))
                    else:
                        Overnight_index += 1

                except (NoSuchElementException, StaleElementReferenceException, IndexError,) as e:
                    #print('No more overnight scan')
                    break
                #print('Overnight ', Overnight_variable_no, ' is done on ', Overnight_var[f'Overnight_{Overnight_variable_no}'])
                cell = worksheet.cell(row=row_to_use, column=overnight_column)
                cell.value = Overnight_var[f'Overnight_{Overnight_variable_no}']
                cell.style = date_format

                # Automatically add 1 to counter and variable no.
                Overnight_variable_no += 1
                Overnight_counter += 1

        # _________________________________________________________#
        # Delivery section:
        # _________________________________________________________#
        try:
            # Delivery_Sign_xpath = (f"//div[@class='el-table__body-wrapper is-scrolling-none']//tbody//"
            #                        f"tr[.//td[contains(., 'Delivery Signature')]and .//td[contains(., {row['Delivery DP']})]]")
            Delivery_Sign_xpath = (f"//div[@class='el-table__body-wrapper is-scrolling-none']//tr[.//td[contains(., 'Delivery Signature')] "
                                   f"and .//td[contains(., '{row['Delivery DP']}')]]")

            Delivery_Sign_elements = driver.find_elements(By.XPATH, Delivery_Sign_xpath)

            #print(len(Delivery_Sign_elements))

            time_xpath = "./td[3]//span"

            Delivery_Sign = datetime.strptime((Delivery_Sign_elements[0].find_element(By.XPATH, time_xpath).text),
                                              '%Y-%m-%d %H:%M:%S')

            cell = worksheet.cell(row=row_to_use, column=13)
            cell.value = Delivery_Sign
            cell.style = date_format
        except (NoSuchElementException, StaleElementReferenceException, IndexError,) as e:
            # print('Return Reg passed')
            pass

        # _________________________________________________________#
        # ATS section:
        # _________________________________________________________#

        try:
            # Delivery_Sign_xpath = (f"//div[@class='el-table__body-wrapper is-scrolling-none']//tbody//"
            #                        f"tr[.//td[contains(., 'Delivery Signature')]and .//td[contains(., {row['Delivery DP']})]]")
            ATS_xpath = (f"//div[@class='el-table__body-wrapper is-scrolling-none']//tr[.//td[contains(., 'Abnormal Termination')] "
                                   f"and .//td[contains(., '{row['Delivery DP']}')]]")

            ATS_elements = driver.find_elements(By.XPATH, ATS_xpath)

            #print(len(Delivery_Sign_elements))

            time_xpath = "./td[3]//span"

            ATS = datetime.strptime((ATS_elements[0].find_element(By.XPATH, time_xpath).text),
                                              '%Y-%m-%d %H:%M:%S')

            cell = worksheet.cell(row=row_to_use, column=14)
            cell.value = ATS
            cell.style = date_format
        except (NoSuchElementException, StaleElementReferenceException, IndexError,) as e:
            # print('Return Reg passed')
            pass

        # add one to row
        row_to_use += 1

    # End of loop ILoc reassignment
    ILoc_Start = ILoc_End + 1
    # on the first loop when counter is zero, use different sum for adding.
    if Counter == 0:
        ILoc_End += 1001
    else:
        ILoc_End += 1000

    # Add 1000 to counter for progress tracking
    Counter+=1000

    # Clear out Parcel Tracking list by clicking on 'empty' button
    empty_but_xpath = "//button[contains(@class, 'el-button b_reset second-btn el-button--small') and .//i[contains(@class, 'el-icon-refresh-right')] and .//span[text()='Empty']]"
    driver.find_element(By.XPATH, empty_but_xpath).click()

#

# Save file
# file name
xlsx_file.save(xlsx_path) # save the file once done writing

# Delete tmp txt file
