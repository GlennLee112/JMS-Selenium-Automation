# Main modules
from datetime import date, datetime
import os
import glob
from datetime import time
import openpyxl as opx
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from openpyxl import Workbook, load_workbook
from xlsx2csv import Xlsx2csv
from pycnnum import num2cn
import datetime as dt
from functools import reduce
import statistics
import shutil
import excel2img as e2i
import shutil

# Feishu BOT module
import hashlib
import base64
import hmac
import requests
import json

# Universal headcount function
# Create a separate list of employee headcount by timeframe

def Headcount(dfcsv,d_type='D'):
    
    time=['10:00:00','12:00:00','14:00:00','16:00:00','18:00:00']
    
    named_dataframes = {}  # Dictionary to store named dataframes
    mode=dfcsv['Delivery Time'].dt.date.mode().iloc[0] # Setup mode for filtering purpose; choose the first on the list with iloc
    
    # Step 0: setting up Delivery or signature, as well as list for column rearranging
    if d_type=='D':
        d_type='Delivery'
        c_type='Delivery Time'
        order=['Dispatcher No.','DP','Employee Type','Delivery before 10AM','Delivery before 12PM','Delivery before 02PM','Delivery before 04PM','Delivery before 06PM']
        dfl=['d10','d12','d14','d16','d18']
        
    elif d_type=='S':
        d_type='Signature'
        c_type='Delivery Signature'
        order=['Dispatcher No.','DP','Employee Type','Signature before 10AM','Signature before 12PM','Signature before 02PM','Signature before 04PM','Signature before 06PM']
        dfl=['ds10','ds12','ds14','ds16','ds18']
            
    # Step 1: Create empty dataframe
    for name in dfl:
        named_dataframes[name] = pd.DataFrame()
    
    # Step 2: Filter and group data for each time interval
    for t,name in zip(time,dfl):
        filtered_df = dfcsv[(dfcsv[c_type].dt.time <= pd.to_datetime(t).time()) & (dfcsv['Delivery Time'].dt.date == mode)]
        filtered_named_df = filtered_df[['Dispatcher No.', c_type , 'Employee Type']].copy()
        grouped_df = filtered_named_df.groupby('Dispatcher No.').size().reset_index(name=d_type+' before '+pd.to_datetime(t).time().strftime('%I%p'))
        named_dataframes[name] = grouped_df

    # Step 4: append and combine dataframe into one
    dfl=[] #Empty list
    for name, df in named_dataframes.items():
        globals()[name] = df
        dfl.append(df)

    # Step 5: create DP and employee type columns for appended data (with time interval)
    dfdl=reduce(lambda  left,right: pd.merge(left,right,on=['Dispatcher No.'],how='outer'), dfl).fillna(0)
    dfdl['DP']=dfdl['Dispatcher No.'].apply(lambda x:x[:6] if len(x)==10 else x[:7])
    dfdl['Employee Type']=dfdl.apply(lambda row:row['Dispatcher No.'].replace(str(row['DP']),""),axis=1).apply(lambda x:x[0:1]).astype(int)
    ## rearrange columns
    dfdl=dfdl.reindex(columns=order)
    
    return dfdl
# Title creation - for pasting in Feishu chat box

def Feishu_Title():
    # Section 1 -- setting up
    ## Part 0 - get current time
    now=datetime.now() # now_h = now in hours (24h format)
    now_h=int(now.strftime("%H"))
    
    ## Part 1 - cn title
    t1='柔佛人员出勤报表' 
    
    ## Part 2 - cn date
    t2=now.strftime("%Y-%m-%d")
    
    ## Part 3 - cn day (e.g.:星期一，星期日)
    ### Use for t3; handling num2cn's bug in converting 1 to '一' and convert 7 to '日'.
    try:
        t3n=num2cn(now.isoweekday())
        if t3n=='七':
            t3n='日'
        else:
            t3n
    except IndexError:
        t3n='一'    
    t3='星期'+t3n
    
    ## Part 4 - en title
    t4='Johor Active Delivery Staff'
    ## Part 5 - en date (same as cn)
    t5=t2
    ## Part 6 - en day (e.g.: Tuesday, wednesday)
    t6=now.strftime("%A")
    
    ## Part 7 - en & cn time (10AM,2PM, etc.)
    match now_h:
        case now_h if now_h>=10 and now_h<12:
            t7=time(10,0,0).strftime('%I%p')
        case now_h if now_h>=12 and now_h<14:
            t7=time(12,0,0).strftime('%I%p')
        case now_h if now_h>=14 and now_h<16:
            t7=time(14,0,0).strftime('%I%p')
        case now_h if now_h>=16 and now_h<18:
            t7=time(16,0,0).strftime('%I%p')
        case now_h if now_h>=18:
            t7=time(18,0,0).strftime('%I%p')
            
    # Section 2 -- Joining title
    ## By part
    join_title1=t1+' ('+t2+', '+t3+', '+t7+')'
    join_title2=t4+' ('+t5+', '+t6+', '+t7+')'
    ## Joining
    l=[join_title1,join_title2]
    title = '\n'.join(l)
    
    return title

# Read the converted CSV file for delivery data
def Delivery_Read(old='N'):
    path=os.getcwd()
    # Read original xlsx file and convert to csv; conversion necessary due to long load time of xlsx file for large data
    if old =='N':
        delivery_folder=os.path.join(path,"Delivery Monitoring")
        delivery_xlsx=os.path.join(path,"Delivery Monitoring","*.xlsx")
        delivery_csv=os.path.join(path,"Delivery Monitoring","*.csv")
    elif old !='N':
        delivery_folder=os.path.join(path,"Delivery Monitoring","Old")
        delivery_xlsx=os.path.join(path,"Delivery Monitoring","Old","*.xlsx")
        delivery_csv=os.path.join(path,"Delivery Monitoring","Old","*.csv")
    
    # Read Excel file and convert into CSV
    delivery_xlsx_file=max(glob.glob(delivery_xlsx),key=os.path.getctime)
    delivery_filename=max(os.path.splitext(os.path.basename(delivery_xlsx_file)))
    Xlsx2csv(delivery_xlsx_file,dateformat='#####.#####').convert(os.path.join(delivery_folder,delivery_filename+'.csv'))
    
    # Read delivery csv data and load to dataframe
    delivery_csv_file=max(glob.glob(delivery_csv))
    delivery_dataframe=pd.read_csv(delivery_csv_file,usecols=['Dispatcher No.','Delivery Time','Delivery Signature','DP No. | Delivery'])
    delivery_dataframe=delivery_dataframe[['DP No. | Delivery','Dispatcher No.','Delivery Time','Delivery Signature']]
    
    return delivery_dataframe

# Read downloaded Arrival file from folder
def Arrival_Read(old='N'):
    path=os.getcwd()
    # Read and load xlsx arrival file to dataframe
    #arrival_folder=os.path.join(path,"Arrival Monitoring")
    if old=='N':
        arrival_xlsx=os.path.join(path,"Arrival Monitoring","*.xlsx")
    elif old!='N':
        arrival_xlsx=os.path.join(path,"Arrival Monitoring","Old","*.xlsx")
    arrival_xlsx_file=max(glob.glob(arrival_xlsx),key=os.path.getctime)
    arrival_dataframe=pd.read_excel(arrival_xlsx_file,usecols=['Operating Station','Qty | Arrived'])
    
    # Delimit last space (' ') and extract value and combine with 'JHR' to get outlet code
    ## reference: https://stackoverflow.com/questions/52139008/pandas-split-string-on-last-occurrence
    arrival_dataframe['DP Code']='JHR'+arrival_dataframe['Operating Station'].str.rsplit(" ",n=1).str[-1]
    
    ## reference: https://stackoverflow.com/questions/13411544/delete-a-column-from-a-pandas-dataframe
    arrival_dataframe.drop('Operating Station',axis=1,inplace=True) # Remember to set inplace = True; this enable dropping column without the need to reassign
    arrival_dataframe=arrival_dataframe[['DP Code','Qty | Arrived']]
                                  
    return arrival_dataframe

# Read downloaded departure file from folder
def Departure_Read():
    path=os.getcwd()
    # Read and load xlsx arrival file to dataframe
    #arrival_folder=os.path.join(path,"Arrival Monitoring")
    Departure_xlsx=os.path.join(path,"Departure Monitoring","*.xlsx")
    Departure_xlsx_file=max(glob.glob(Departure_xlsx),key=os.path.getctime)
    Departure_dataframe=pd.read_excel(Departure_xlsx_file,usecols=['操作站点编号','Next Station Code','Total | Departure'])
    Departure_dataframe=Departure_dataframe[['操作站点编号','Next Station Code','Total | Departure']]
                                  
    return Departure_dataframe

# 4. Data processing
## 4.1 Base data preparation part (creating dfcsv)
### 4.1.1 Fill in empty delivery signature time with blank ('') 
def Delivery_Cleaning(df):
    df=df
    df[["Delivery Signature"]]=df[["Delivery Signature"]].fillna('') 
        ### 4.1.2 Filter out dispatcher no. that doesn't contain 'JHR' at the start
    df=df[df["Dispatcher No."].str.contains('JHR')]
        ### 4.1.3 Convert Delivery and delivery signature time to proper datetime format.
    df['Delivery Time']=pd.to_datetime(df['Delivery Time'])
    df['Delivery Signature']=pd.to_datetime(df['Delivery Signature'])
        ### 4.1.4 Add Employee Type column
    # Reference - https://stackoverflow.com/questions/54892624/remove-substring-from-column-based-on-another-column
    df['Employee Type']=df.apply(lambda row:row['Dispatcher No.'].replace(str(row['DP No. | Delivery']),""),axis=1).apply(lambda x:x[0:1])
    
    #Old Backup - df['Employee Type']=df['Dispatcher No.'].apply(lambda x:x[6:7] if len(x)==10 else x[7:8]) # Lamba to get the 7th number in dispatcher no. if lenght of no. is 10, else get the 8th number.
        
        ### 4.1.5 Convert Employee Type to int
    df['Employee Type']=df['Employee Type'].astype('int') # Convert the result from lambda into integer (when first convert it is in string format)
        ### 4.1.7 Replace empty delivery & signature entry with NaT into blank ('')
    df.replace({pd.NaT: ''}, inplace=True)
    df=df[['DP No. | Delivery','Dispatcher No.','Employee Type','Delivery Time','Delivery Signature']]
    return df

## 5. File Select function
    ## If 10am time frame, use template.
def File_Select():
    path=os.getcwd()
    # Set current time (hours) for checking
    Now=datetime.now()
    Now_h=int(Now.strftime("%H"))
    
    # Logic gate - if between 10 and 12, copy template and rename file
    if Now_h >= 10 and Now_h < 12:
        
        # Partial file grab; fetch files with '柔佛人员出勤报表' name into a list
        file_use=glob.glob('柔佛人员出勤报表*.xlsx') 
        
        # Grab first file of the file list
        xlsxfile=os.path.join(path,file_use[0]) 
        
        # Copy template to 'export' folder
        shutil.copy(xlsxfile,os.path.join(path,"Export"))
        ef=os.path.join(path,"Export")
        
        # Create list of files in 'Export' folder & select the first file
        efl=glob.glob(os.path.join(path,"Export","*.xlsx"))
        xlsxfile=os.path.join(ef,efl[0])
        
        # Date to rename file with
        File_date=Now.strftime('%m%d')
        
        # Set up directory
        directory=os.path.dirname(xlsxfile)
        
        # Renaming
        os.rename(xlsxfile,os.path.join(directory,"柔佛人员出勤报表"+File_date+".xlsx"))
        
        # Reselecting new file to use
        ef=os.path.join(path,"Export","*.xlsx")
        #efl=os.listdir(os.path.join(path,"Export"))
        xlsxfile=max(glob.glob(ef),key=os.path.getctime)
        
    else:
        # Select
        ef=os.path.join(path,"Export","*.xlsx")
        #efl=os.listdir(os.path.join(path,"Export"))
        xlsxfile=max(glob.glob(ef),key=os.path.getctime)
    
    return xlsxfile

## 6. Excel writing
## Universal Excel Writing
def Excel_Write(XLSX,DF,Sheets_No,Row=1,Column=1):
    
    wb=load_workbook(XLSX)
    ws=wb.worksheets[Sheets_No]
    
    for idx, row in DF.iterrows():
        for col in range(len(row)):
            column_letter = opx.utils.get_column_letter(Column+col)
            ws[f'{column_letter}{idx+Row+1}'] = DF.iloc[idx, col]
            
    wb.save(XLSX)
    
# Range maker for use by column list

def Range_maker_number(start,end):  # Setup numerical range maker for use to iterate column and cell ranges
    List=[start]   
    while end!=start:
        start+=1
        List.append(start)
        
    return List

## Hide Column

def Hide_Col(file,Range,Sheets_No):
    
    wb=load_workbook(file)
    ws=wb.worksheets[Sheets_No]
    
    for r in Range:
        r=opx.utils.get_column_letter(r)
        ws.column_dimensions[r].hidden=True
        
    wb.save(file)

## Unhide Column

def Unhide_Col(file,Range,Sheets_No):
    
    wb=load_workbook(file)
    ws=wb.worksheets[Sheets_No]
    
    min_r=opx.utils.get_column_letter(min(Range))
    max_r=opx.utils.get_column_letter(max(Range))
    
    ws.column_dimensions.group(start=min_r, end=max_r, hidden=False)
        
    wb.save(file)

## Overnight Writing

# JMS interface

#------------------------------------------#

## Business indicator > Operation > General DP Overnight Montoring
## Date: 1 month range from the day before (i.e.: on 17th July, insert date starting  16th June 
##       to 16th July)

# Excel requirement
## Arrival Time -- remove arrival of today (i.e.: 17th July, remove all entry with 17th July arrival)
##              -- remove blank arrival as well
## Return Parcel -- remove all Return parcel with 'Y'
    
def Overnight_Read():
    """
    Reading Downloaded General DP OVernight File from JMS
    
    Involving CVS conversion and then read to circumvent long Excel load time
    
    """
    # Conversion to CSV and then read
    
    ## Conversion part
    path=os.getcwd()
    overnight_folder=os.path.join(path,"General DP Overnight")
    overnight_xlsx=os.path.join(path,"General DP Overnight","*.xlsx")
    overnight_xlsx_file=max(glob.glob(overnight_xlsx),key=os.path.getctime)
    overnight_filename=max(os.path.splitext(os.path.basename(overnight_xlsx_file)))
    Xlsx2csv(overnight_xlsx_file,dateformat='#####.#####').convert(os.path.join(overnight_folder,overnight_filename+'.csv'))
    
    ## Read part
    overnight_csv=os.path.join(path,"General DP Overnight","*.csv")
    overnight_csv_file=max(glob.glob(overnight_csv))
    overnight_dataframe=pd.read_csv(overnight_csv_file,usecols=['Scanning Type | Last','Arrival Time','Scanning DP No. | Last','Scanning Time | Last','Return Parcel'])
                                  
    return overnight_dataframe

def Overnight_Cleaning(df):
    # Set up df
    ONDF=df
    # 1. Datetime conversion; convert to datetime and then date
    ONDF[['Arrival Time','Scanning Time | Last']] = ONDF[['Arrival Time','Scanning Time | Last']].apply(pd.to_datetime, errors='coerce')
    ONDF['Scanning Time | Last']=ONDF['Scanning Time | Last'].dt.date
    ONDF['Arrival Time']=ONDF['Arrival Time'].dt.date
    
    # 2. Data cleaning - replace NA date time with blank for ease of filtering
    ONDF['Arrival Time']=ONDF['Arrival Time'].fillna('')
    ONDF=ONDF[(ONDF['Arrival Time']!="")]
    
    # 3. Data filtering and processing
    ## Todays date, get max date
    Max_Date=max(ONDF['Arrival Time'])
    ## Filter by condition
    ### Filter out
    ONDF=ONDF[(~ONDF['Scanning DP No. | Last'].str.contains("JHR0",regex=True)==True) & # Unnecessary DP; note the tilde (~) operator for logic inversion
              (ONDF['Arrival Time']<Max_Date) & # Today Arrival removal
              (ONDF['Return Parcel']!='Y')] # Filter out all return parcel
    
    ## Grouping and summarizing data into a short summarized table
    ONDF=ONDF.groupby('Scanning DP No. | Last').size().reset_index()
    ONDF.reset_index(inplace=True,drop=True)
    ONDF.columns.values[1]='Overnight Volume'
    
    return ONDF

def Yesterday_Overnight():
    ## Set up file path
    path=os.getcwd() # base directory
    ef=os.path.join(path,"Export","*.xlsx") # file list with .xlsx extension
    xlsxfile=max(glob.glob(ef),key=os.path.getctime) # Get the latest file
    
    ## Reading and writing overnight data
    Old_Overnight=pd.read_excel(xlsxfile,sheet_name='Overnight')
    Old_Overnight=Old_Overnight[['Today Overnight_DP','Today Overnight']]
    Old_Overnight.to_excel(os.path.join(path,'Yesterday Overnight','Yesterday Overnight.xlsx'),index=False)

def Yesterday_Overnight_Read():
    ## Reading yesterday overnight data writen from yesterday's report file
    path=os.getcwd()

    Y_Overnight_xlsx=os.path.join(path,"Yesterday Overnight","*.xlsx")
    Y_Overnight_xlsx_file=max(glob.glob(Y_Overnight_xlsx),key=os.path.getctime)
    Y_Overnight_dataframe=pd.read_excel(Y_Overnight_xlsx_file)
                                  
    return Y_Overnight_dataframe

def Image(File,Col):
## Solution adopted -- excel2img
# Reference: https://stackoverflow.com/questions/37180097/copy-image-from-excel-and-save-it-using-python
    path = os.getcwd()
    File_date=datetime.now().strftime('%m%d')
    FileName=os.path.splitext(os.path.basename(File))[0]
    e2i_File=File
    e2i_File_png=os.path.join(path,'Export','IMG',FileName+'.PNG')
    ## Column
    e2i_col=opx.utils.get_column_letter(min(Col)-1)

    cell_use=f'T!'+'B4'+':'+e2i_col+'63'
    ## Execute export image
    e2i.export_img(e2i_File,e2i_File_png,"",cell_use)
    
# Nuke folder -- Clean out all folder(s) specified by user

## https://stackoverflow.com/questions/1995373/deleting-all-files-in-a-directory-with-python
## https://www.tutorialspoint.com/How-to-delete-all-files-in-a-directory-with-Python
## https://stackoverflow.com/questions/22812785/use-endswith-with-multiple-extensions

## Changes: get the full list of flies in provided link
def nuke_folder(folderlist):
    os.getcwd()
    # Loop iterating through folder
    for f1 in folderlist:
        # Go into each folder and extract all file(s)
        for files in os.walk(f1):
            for fn in files[2]: # Set up index of [2] to specify looking for the file list tuple generated with os.walk 
                # (for more details: https://stackoverflow.com/questions/34556433/what-are-the-empty-lists-in-dirname-of-os-walk)
                os.remove(os.path.join(path,f1,fn))


def Format_Column(file, Sheet, hide_col):
    wb = load_workbook(file)

    if isinstance(Sheet, int):
        ws = wb.worksheets[Sheet]
    elif isinstance(Sheet, str):
        ws = wb[Sheet]

    # list for comparison
    List_8 = ['全职\nFull time', '兼职\nPart time', '第三方\nThird party']  # List for item with width of 8
    List_10 = ['全职\nFull time', '周环比\nCompare to Last Week', '出仓量\nDelivering Volume',
               '签收量\nSignature Volume']  # List for item with width of 10
    List_11 = ['人员\nManpower']  # List for item with width of 11
    List_13 = ['出仓率\nDelivery Rate\n(Target: 30%)', '出仓率\nDelivery Rate\n(Target: 50%)',
               '出仓率\nDelivery Rate\n(Target: 85%)', '出仓率\nDelivery Rate\n(Target: 90%)',
               '出仓率\nDelivery Rate\n(Target: 95%)']  # List for item with width of 13

    Col_to_format = Range_maker_number(14, min(hide_col))

    for Col_no in Col_to_format:

        Col_A = opx.utils.get_column_letter(Col_no)

        value = ws[Col_A + '7'].value

        if value in List_8:
            ws.column_dimensions[str(Col_A)].width = 8

        if value in List_10:
            ws.column_dimensions[str(Col_A)].width = 10

        if value in List_11:
            ws.column_dimensions[str(Col_A)].width = 11

        if value in List_13:
            ws.column_dimensions[str(Col_A)].width = 13

    wb.save(file)


def gen_sign(timestamp, secret):
    # Splicing timestamp and secret
    string_to_sign = '{}\n{}'.format(timestamp, secret)
    hmac_code = hmac.new(string_to_sign.encode("utf-8"), digestmod=hashlib.sha256).digest()
    # Perform base64 processing on the result
    sign = base64.b64encode(hmac_code).decode('utf-8')
    return sign

def Excel_Date(file,date):
    # load workbook and sheet
    wb = load_workbook(file)
    ws = wb.worksheets[0]

    # setup date and cell to write
    ws['B1'] = date

    # save
    wb.save(file)

def Archive(file):
    # Set up base path
    path = os.getcwd()

    # Destination folder
    destination = os.path.join(path,"Archive")

    # Old file path
    old_file = file

    now_time = datetime.now()
    now_year_month = now_time.strftime("%Y.%m")
    now_date = now_time.strftime("%m%d")

    New_Folder = os.path.join(destination, now_year_month)
    #New_File = os.path.join(destination, now_year_month,"柔佛人员出勤报表"+now_date+".xlsx")

    if os.path.exists(New_Folder):
        pass
    elif not os.path.exists(New_Folder):
        os.makedirs(New_Folder)

    shutil.move(old_file, os.path.join(destination,now_year_month,"柔佛人员出勤报表"+now_date+".xlsx"))

#def Archive_file():
    # At the end of the day, copy file into archive

# Feishu access and upload files

## Solution 1: PyAutoGui (on hold)
## G: Ease
## B: Unstable

## Solution 2: PyWinAuto (on hold)
## G: moderately integrated solution
## B: not yet figure out how to do

## Solution 3: Feishu Lark API / BOT (SUCCESS WTF)
## G: most integrated solution and open up to future iteration
## B: utmost difficult solution; high lieklihood of not working without external help

    # Reference:
        # https://open.feishu.cn/document/client-docs/bot-v3/bot-overview
        # https://stackoverflow.com/questions/58796540/python-post-request-to-api (HTTP post request)
        # https://open.feishu.cn/document/client-docs/bot-v3/add-custom-bot (security settings)